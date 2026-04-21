from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "output"

ESTABLECIMIENTOS = Path(
    os.environ.get(
        "MAESTRO_ESTABLECIMIENTOS_PATH",
        ROOT / "data" / "establecimientos_20260406_oficial.csv",
    )
)

NUMERADOR = OUTPUT / "numerador_p12_b1_mamografia_rm_2025_resumen_establecimiento.csv"
DENOMINADOR = OUTPUT / "denominador_poblacion_inscrita_validada_mujeres_50_69_rm_base_pago_2025.csv"


def code_text(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace({"nan": "", "None": ""})
    )


def load_establecimientos() -> pd.DataFrame:
    cols = [
        "EstablecimientoCodigo",
        "EstablecimientoCodigoMadreNuevo",
        "SeremiSaludGlosa_ServicioDeSaludGlosa",
        "TipoEstablecimientoGlosa",
        "EstablecimientoGlosa",
        "DependenciaAdministrativa",
        "ComunaGlosa",
        "EstadoFuncionamiento",
    ]
    df = pd.read_csv(ESTABLECIMIENTOS, sep=";", dtype=str, usecols=cols)
    for col in ["EstablecimientoCodigo", "EstablecimientoCodigoMadreNuevo"]:
        df[col] = code_text(df[col])
    return df.drop_duplicates("EstablecimientoCodigo")


def add_master_fields(df: pd.DataFrame, estab: pd.DataFrame) -> pd.DataFrame:
    out = df.merge(
        estab,
        left_on="IdEstablecimiento",
        right_on="EstablecimientoCodigo",
        how="left",
    ).drop(columns=["EstablecimientoCodigo"])
    return out.rename(
        columns={
            "EstablecimientoCodigoMadreNuevo": "codigo_madre_master",
            "SeremiSaludGlosa_ServicioDeSaludGlosa": "servicio_salud_master",
            "TipoEstablecimientoGlosa": "tipo_establecimiento_master",
            "EstablecimientoGlosa": "establecimiento_master",
            "DependenciaAdministrativa": "dependencia_master",
            "ComunaGlosa": "comuna_master",
            "EstadoFuncionamiento": "estado_funcionamiento_master",
        }
    )


def build_establecimiento(numerador: pd.DataFrame, denom: pd.DataFrame) -> pd.DataFrame:
    denom_cols = [
        "IdEstablecimiento",
        "servicio_salud_denominador",
        "dependencia_denominador",
        "comuna_denominador",
        "establecimiento_denominador",
        "denominador_poblacion_inscrita_validada_mujeres_50_69",
    ]
    out = numerador.merge(denom[denom_cols], on="IdEstablecimiento", how="left")
    out = add_master_fields(out, load_establecimientos())
    out["denominador_disponible"] = out["denominador_poblacion_inscrita_validada_mujeres_50_69"].notna()
    out["cobertura_mamografia_pct"] = (
        out["numerador_mujeres_50_69"]
        / out["denominador_poblacion_inscrita_validada_mujeres_50_69"]
        * 100
    ).where(out["denominador_poblacion_inscrita_validada_mujeres_50_69"].gt(0))
    out["advertencia_cobertura_establecimiento"] = ""
    out.loc[
        ~out["denominador_disponible"],
        "advertencia_cobertura_establecimiento",
    ] = "Sin población inscrita y validada directa; no usar como cobertura establecimiento"
    return out.sort_values(["Mes", "IdServicio", "IdComuna", "IdEstablecimiento"])


def build_comuna(cob_estab: pd.DataFrame, denom: pd.DataFrame) -> pd.DataFrame:
    num_com = (
        cob_estab.groupby(["Ano", "Mes", "IdRegion", "IdComuna"], as_index=False)[
            "numerador_mujeres_50_69"
        ]
        .sum()
    )
    num_com["IdComuna"] = code_text(num_com["IdComuna"])
    denom["IdComuna_master"] = code_text(denom["IdComuna_master"])
    denom_com = (
        denom.dropna(subset=["IdComuna_master"])
        .groupby("IdComuna_master", as_index=False)["denominador_poblacion_inscrita_validada_mujeres_50_69"]
        .sum()
        .rename(columns={"IdComuna_master": "IdComuna"})
    )
    names = (
        denom.dropna(subset=["IdComuna_master"])
        .groupby("IdComuna_master", as_index=False)["comuna_master"]
        .first()
        .rename(columns={"IdComuna_master": "IdComuna"})
    )
    out = num_com.merge(denom_com, on="IdComuna", how="left").merge(
        names, on="IdComuna", how="left"
    )
    out["cobertura_mamografia_pct"] = (
        out["numerador_mujeres_50_69"]
        / out["denominador_poblacion_inscrita_validada_mujeres_50_69"]
        * 100
    ).where(out["denominador_poblacion_inscrita_validada_mujeres_50_69"].gt(0))
    return out.sort_values(["Mes", "IdComuna"])


def build_control(cob_estab: pd.DataFrame) -> pd.DataFrame:
    sin_denominador = cob_estab[~cob_estab["denominador_disponible"]].copy()

    controles = []
    cols = [
        "tipo_control",
        "Mes",
        "IdServicio",
        "IdComuna",
        "IdEstablecimiento",
        "numerador_mujeres_50_69",
        "denominador_poblacion_inscrita_validada_mujeres_50_69",
        "cobertura_mamografia_pct",
        "codigo_madre_master",
        "servicio_salud_master",
        "comuna_master",
        "establecimiento_master",
        "tipo_establecimiento_master",
        "estado_funcionamiento_master",
    ]

    if not sin_denominador.empty:
        sin_denominador["tipo_control"] = "REM con numerador P12 sin población inscrita y validada directa"
        controles.append(sin_denominador[cols])

    if not controles:
        return pd.DataFrame({"tipo_control": ["Sin observaciones"]})
    return pd.concat(controles, ignore_index=True)


def metadata_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("anio", "2025"),
            ("region", "13 - Region Metropolitana"),
            ("numerador", "REM-P12, Seccion B1, Col01, codigos P1220030/P1207030/P1207040/P1207050"),
            ("denominador", "Población inscrita y validada mujeres 50-69, inscritos 2024, base de pago 2025"),
            ("alias_denominador_aplicado", "311001 -> 201674, Cesfam El Abrazo Dr. Salvador Allende"),
            ("corte_principal", "Mes 12 del REM-P12 2025"),
            ("formula", "numerador_mujeres_50_69 / denominador_poblacion_inscrita_validada_mujeres_50_69 * 100"),
            ("uso_recomendado", "Usar cobertura comunal como salida principal; la cobertura por establecimiento es referencial/control."),
            ("nota", "No se usa REM-A29 porque registra examenes/procedimientos, no personas unicas con vigencia."),
        ],
        columns=["campo", "valor"],
    )


def safe_to_csv(df: pd.DataFrame, path: Path) -> Path:
    try:
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_actualizado{path.suffix}")
        df.to_csv(fallback, index=False, encoding="utf-8-sig")
        return fallback


def format_excel(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 45)
    wb.save(path)


def write_visualizacion_excel(
    path: Path,
    cob_estab: pd.DataFrame,
    cob_comuna: pd.DataFrame,
    control: pd.DataFrame,
    meta: pd.DataFrame,
) -> Path:
    comuna_cols = [
        "Ano",
        "Mes",
        "IdRegion",
        "IdComuna",
        "comuna_master",
        "numerador_mujeres_50_69",
        "denominador_poblacion_inscrita_validada_mujeres_50_69",
        "cobertura_mamografia_pct",
    ]
    estab_cols = [
        "Ano",
        "Mes",
        "IdServicio",
        "IdComuna",
        "comuna_master",
        "IdEstablecimiento",
        "establecimiento_master",
        "tipo_establecimiento_master",
        "50_54",
        "55_59",
        "60_64",
        "65_69",
        "numerador_mujeres_50_69",
        "denominador_poblacion_inscrita_validada_mujeres_50_69",
        "cobertura_mamografia_pct",
        "advertencia_cobertura_establecimiento",
    ]

    def write(path_to_write: Path) -> Path:
        with pd.ExcelWriter(path_to_write, engine="openpyxl") as writer:
            cob_comuna[comuna_cols].to_excel(writer, sheet_name="comuna_principal", index=False)
            cob_estab[estab_cols].to_excel(writer, sheet_name="establecimiento_control", index=False)
            control.to_excel(writer, sheet_name="control_calidad", index=False)
            meta.to_excel(writer, sheet_name="metodologia", index=False)
        format_excel(path_to_write)
        return path_to_write

    try:
        return write(path)
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_actualizado{path.suffix}")
        return write(fallback)


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    if not NUMERADOR.exists():
        raise FileNotFoundError(f"Falta numerador. Ejecuta primero: {ROOT / '01_extraer_numerador_p12.py'}")
    if not DENOMINADOR.exists():
        raise FileNotFoundError(f"Falta denominador. Ejecuta primero: {ROOT / '02_calcular_denominador_poblacion_inscrita_validada.py'}")

    numerador = pd.read_csv(NUMERADOR, dtype=str)
    denom = pd.read_csv(DENOMINADOR, dtype=str)
    for df in [numerador, denom]:
        df["IdEstablecimiento"] = code_text(df["IdEstablecimiento"])

    int_cols = ["Ano", "Mes", "IdRegion", "IdServicio", "IdComuna", "50_54", "55_59", "60_64", "65_69", "numerador_mujeres_50_69"]
    for col in int_cols:
        numerador[col] = pd.to_numeric(numerador[col], errors="coerce").astype("int64")
    denom["denominador_poblacion_inscrita_validada_mujeres_50_69"] = pd.to_numeric(
        denom["denominador_poblacion_inscrita_validada_mujeres_50_69"], errors="coerce"
    )

    cob_estab = build_establecimiento(numerador, denom)
    cob_comuna = build_comuna(cob_estab, denom)
    control = build_control(cob_estab)
    meta = metadata_frame()

    paths = {
        "cobertura_mamografia_establecimiento_rm_2025.csv": cob_estab,
        "cobertura_mamografia_comuna_rm_2025.csv": cob_comuna,
        "control_calidad_mamografia_rm_2025.csv": control,
        "metadata_mamografia_rm_2025.csv": meta,
    }
    written_paths = []
    for filename, df in paths.items():
        written_paths.append(safe_to_csv(df, OUTPUT / filename))

    def write_workbook(path: Path) -> Path:
        try:
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                cob_comuna.to_excel(writer, sheet_name="comuna_dic", index=False)
                cob_estab.to_excel(writer, sheet_name="establecimiento_control", index=False)
                denom.to_excel(writer, sheet_name="denominador", index=False)
                control.to_excel(writer, sheet_name="control", index=False)
                meta.to_excel(writer, sheet_name="metodologia", index=False)
            format_excel(path)
            return path
        except PermissionError:
            fallback = path.with_name(f"{path.stem}_actualizado{path.suffix}")
            with pd.ExcelWriter(fallback, engine="openpyxl") as writer:
                cob_comuna.to_excel(writer, sheet_name="comuna_dic", index=False)
                cob_estab.to_excel(writer, sheet_name="establecimiento_control", index=False)
                denom.to_excel(writer, sheet_name="denominador", index=False)
                control.to_excel(writer, sheet_name="control", index=False)
                meta.to_excel(writer, sheet_name="metodologia", index=False)
            format_excel(fallback)
            return fallback

    workbook = OUTPUT / "cobertura_mamografia_rm_2025.xlsx"
    workbook_written = write_workbook(workbook)
    visualizacion = OUTPUT / "visualizacion_paralela_mamografia_dic_2025.xlsx"
    visualizacion_written = write_visualizacion_excel(
        visualizacion, cob_estab, cob_comuna, control, meta
    )

    dec = cob_comuna[cob_comuna["Mes"].eq(12)]
    cobertura_rm = (
        dec["numerador_mujeres_50_69"].sum()
        / dec["denominador_poblacion_inscrita_validada_mujeres_50_69"].sum()
        * 100
    )
    for path in written_paths:
        print(f"CSV escrito: {path}")
    print(f"Workbook: {workbook_written}")
    print(f"Visualizacion paralela: {visualizacion_written}")
    print(f"Numerador diciembre RM: {dec['numerador_mujeres_50_69'].sum():,}")
    print(f"Denominador diciembre RM: {dec['denominador_poblacion_inscrita_validada_mujeres_50_69'].sum():,.0f}")
    print(f"Cobertura regional diciembre RM: {cobertura_rm:.2f}%")


if __name__ == "__main__":
    main()
